# -*- coding: utf-8 -*-
"""
Created on Thu Dec 14 18:49:59 2023

@author: dmlys
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import time

# Set the database list
db_list = [
    "us", "uk", "br", "ca", "au", "fr", "de", "it", "nl", "es", "in", "ru",
    "jp", "tr", "dk", "mx", "ar", "pl", "be", "ie", "se", "ch", "fi", "hu",
    "no", "il", "sg", "hk", "ae", "sa", "om", "kw", "bh", "lu", "id", "za"
]

# Set the URL and keyword input
url_template = "https://www.semrush.com/analytics/keywordoverview/?q={keyword}&db={db}&date={date}"
keywords_input = input("Enter the list of keywords (one per line):\n").splitlines()

# Validate the database input
db_input = None
while db_input not in db_list:
    if db_input is not None:
        print("Wrong database input. Must be one of the options above.")

    db_input = input("Enter your database (" + str(db_list) + "): ")

# Get the list of dates
dates_input = input("Enter the list of dates (format: YYYYMM, separated by commas):\n").split(",")

print("Ok, Let's go!")

# Create the Excel workbook
wb = openpyxl.Workbook()

# Create the first sheet with the name from db_list + " Volume"
sheet1 = wb.active
sheet1.title = f"{db_input.capitalize()} Volume"

# Set the headers in the first row of sheet1
sheet1.cell(row=1, column=1, value="Keyword")
for i, date in enumerate(dates_input, start=2):
    column_letter = get_column_letter(i)
    sheet1.cell(row=1, column=i, value=date)

# Set the path to the downloaded web driver executable
driver_path = "C:/Users/user/Desktop/chromedriver-win64/chromedriver.exe"  # Replace with the actual path to the web driver executable

# Set the login URL and credentials
login_url = "https://www.semrush.com/login/"
email = "input_your_email_here"
password = "input_your_password_here"

# Create a new instance of the web driver
options = webdriver.ChromeOptions()
options.add_argument('--no-sandbox')  # Add this option if you encounter sandbox-related issues

# Use webdriver.Chrome() without specifying executable_path
driver = webdriver.Chrome(options=options)  # Replace with the appropriate web driver

# Navigate to the login page
driver.get(login_url)

# Find the email and password input fields and enter the credentials
email_input = driver.find_element(By.ID, "email")
password_input = driver.find_element(By.ID, "password")

email_input.send_keys(email)
password_input.send_keys(password)

# Submit the login form
login_button = driver.find_element(By.CSS_SELECTOR, "button[data-ui-name='Button'][data-ga-label='login']")
driver.execute_script("arguments[0].click();", login_button)

# Wait for the login process to complete (you can adjust the time if needed)
driver.implicitly_wait(10)

# Pause the script and wait for user input
input("Please, make sure you are logged in. If not, do it manually and then press Enter")

# Continue with the script execution

# Scraping process for each keyword and date
for keyword in keywords_input:
    # Set the row index for the keyword
    row_sheet1 = sheet1.max_row + 1

    # Fill in the keyword in the first column of sheet1
    sheet1.cell(row=row_sheet1, column=1, value=keyword)

    # Scrape data for each date
    for i, date in enumerate(dates_input, start=2):
        # Construct the URL for the keyword and date
        url = url_template.format(keyword=keyword, db=db_input, date=date)

        # Navigate to the URL
        driver.get(url)

        # Wait for 3 seconds before scraping data
        time.sleep(3)

        # Scrape the data for sheet1
        attempts = 0
        while attempts < 4:
            try:
                element_sheet1 = driver.find_element(By.CSS_SELECTOR, "div.kwo-metrics-data-layout")
                data_sheet1 = element_sheet1.text
                break
            except NoSuchElementException:
                data_sheet1 = "N/A"
                break
            else:
                if attempts == 3:
                    data_sheet1 = ""
                    break
                driver.refresh()
                attempts += 1

        # Write the data to the corresponding cell in sheet1
        sheet1.cell(row=row_sheet1, column=i, value=data_sheet1)
        wb.save("output.xlsx")  # Save the workbook after each value is added

# Save the final version of the Excel workbook
wb.save("output.xlsx")

# Close the web driver
driver.quit()

print("Scraping complete. Excel output file 'output.xlsx' has been generated.")
